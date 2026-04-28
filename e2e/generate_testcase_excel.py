#!/usr/bin/env python3
"""生成 E2E 测试用例矩阵 Excel，包含期望结果基线和自动化覆盖状态"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path(__file__).parent / "testcases.xlsx"

# ===== 样式定义 =====
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
P0_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
P1_FILL = PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")
P2_FILL = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
COVERED_FILL = PatternFill(start_color="D3F9D8", end_color="D3F9D8", fill_type="solid")
NOT_COVERED_FILL = PatternFill(start_color="FFE3E3", end_color="FFE3E3", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFF3BF", end_color="FFF3BF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

PRIORITY_FILLS = {"P0": P0_FILL, "P1": P1_FILL, "P2": P2_FILL}
COVERAGE_FILLS = {"covered": COVERED_FILL, "partial": PARTIAL_FILL, "not_covered": NOT_COVERED_FILL}


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_rows(ws, start_row, end_row, col_count, priority_col=None, coverage_col=None):
    for row in range(start_row, end_row + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = WRAP
        if priority_col:
            p = ws.cell(row=row, column=priority_col).value
            if p in PRIORITY_FILLS:
                ws.cell(row=row, column=priority_col).fill = PRIORITY_FILLS[p]
        if coverage_col:
            c = ws.cell(row=row, column=coverage_col).value
            if c == "已覆盖":
                ws.cell(row=row, column=coverage_col).fill = COVERED_FILL
            elif c == "部分覆盖":
                ws.cell(row=row, column=coverage_col).fill = PARTIAL_FILL
            elif c == "未覆盖":
                ws.cell(row=row, column=coverage_col).fill = NOT_COVERED_FILL


# ===== Sheet 1: 测试用例矩阵 =====
def create_testcase_sheet(wb):
    ws = wb.active
    ws.title = "测试用例"

    headers = ["用例ID", "分类", "场景", "前置条件", "操作步骤", "期望结果（验证点）",
               "Fixture文件", "优先级", "自动化覆盖", "对应测试文件", "备注"]
    ws.append(headers)

    # fmt: (id, category, scenario, precondition, steps, expected, fixture, priority, coverage, test_file, note)
    cases = [
        # ===== 核心管道 =====
        ("C01", "核心管道", "基础脱敏 - xlsx",
         "已创建工作区",
         "1. 导入 sample.xlsx\n2. 等待处理完成\n3. 检查对比视图",
         "1. 高亮数 ≥ 25（5人×5列）\n2. 手机号 13800138000 被高亮\n3. 身份证 110101199001011234 被高亮\n4. 导出文件中对应位置已脱敏",
         "sample.xlsx", "P0", "部分覆盖", "test_basic_desensitize.py",
         "当前只断言 highlights>0，未验证具体项"),

        ("C02", "核心管道", "基础脱敏 - csv",
         "已创建工作区",
         "同 C01，文件换为 sample.csv",
         "同 C01，高亮数 ≥ 25",
         "sample.csv", "P0", "部分覆盖", "test_basic_desensitize.py",
         "同上"),

        ("C03", "核心管道", "基础脱敏 - docx",
         "已创建工作区",
         "同 C01，文件换为 sample.docx",
         "1. 高亮数 ≥ 8\n2. 张三/李四被高亮\n3. 手机号/身份证被高亮",
         "sample.docx", "P0", "部分覆盖", "test_basic_desensitize.py",
         "同上"),

        ("C04", "核心管道", "基础脱敏 - txt",
         "已创建工作区",
         "同 C01，文件换为 sample.txt",
         "1. 高亮数 ≥ 8\n2. 含手机号13800138000\n3. 含身份证110101199001011234\n4. 含邮箱zhangsan@example.com",
         "sample.txt", "P0", "部分覆盖", "test_basic_desensitize.py",
         "同上"),

        ("C05", "核心管道", "基础脱敏 - pdf",
         "已创建工作区，有 pdf fixture",
         "导入 sample.pdf → 等待处理",
         "遮挡区域出现",
         "sample.pdf", "P1", "未覆盖", "", "需先创建 pdf fixture"),

        ("C06", "核心管道", "密码保护文件",
         "已创建工作区",
         "1. 导入 encrypted.xlsx\n2. 弹出密码框\n3. 输入 test123\n4. 等待处理",
         "1. 弹出密码输入框\n2. 输入正确密码后正常识别\n3. 高亮数 > 0",
         "sample_encrypted.xlsx", "P1", "未覆盖", "", "需创建加密 fixture"),

        ("C07", "核心管道", "密码错误重试",
         "已创建工作区",
         "1. 导入 encrypted.xlsx\n2. 输入错误密码\n3. 验证错误提示\n4. 重新输入正确密码",
         "1. 显示错误提示\n2. 可重试\n3. 正确密码后正常处理",
         "sample_encrypted.xlsx", "P2", "未覆盖", "", ""),

        ("C08", "核心管道", "空文件",
         "已创建工作区",
         "导入 empty.xlsx",
         "1. 不崩溃\n2. 高亮数 = 0\n3. 对比视图正常显示",
         "empty.xlsx", "P2", "未覆盖", "", "需创建空 fixture"),

        ("C09", "核心管道", "大文件",
         "已创建工作区",
         "导入 large.csv (1万行)",
         "1. 60秒内完成处理\n2. 高亮数 > 0\n3. 不卡死",
         "large.csv", "P2", "未覆盖", "", "需创建大文件 fixture"),

        # ===== 策略切换 =====
        ("S01", "策略切换", "Mask 策略验证",
         "导入 sample.txt，默认 Mask 策略",
         "1. 等待处理完成\n2. 检查脱敏后文本",
         "1. 手机号脱敏为 138****8000 形式\n2. 身份证脱敏为 1101***********234 形式",
         "sample.txt", "P0", "未覆盖", "", "需检查脱敏后的具体文本"),

        ("S02", "策略切换", "Replace 策略验证",
         "导入 sample.txt",
         "1. 右侧面板切换为 Replace\n2. 检查脱敏后文本",
         "1. 手机号被替换为假手机号（非原号码）\n2. 与 Mask 结果不同",
         "sample.txt", "P0", "未覆盖", "", ""),

        ("S03", "策略切换", "Generalize 策略验证",
         "导入 sample.txt",
         "1. 切换为 Generalize\n2. 检查脱敏后文本",
         "1. 地址被泛化（如保留到市级）\n2. 与 Mask/Replace 结果不同",
         "sample.txt", "P1", "未覆盖", "", ""),

        ("S04", "策略切换", "策略来回切换",
         "导入 sample.txt",
         "1. 默认 Mask → 记录文本A\n2. 切 Replace → 记录文本B\n3. 切回 Mask → 记录文本C",
         "1. A ≠ B\n2. A = C（切回后结果一致）",
         "sample.txt", "P1", "未覆盖", "", ""),

        ("S05", "策略切换", "Replace 风格 Fake/Mou/Ordinal",
         "导入 sample.txt，Replace 策略",
         "分别选 Fake/Mou/Ordinal",
         "Fake: 随机假名\nMou: 某某\nOrdinal: 人员1",
         "sample.txt", "P2", "未覆盖", "", ""),

        ("S06", "策略切换", "Mask 前后缀参数",
         "导入 sample.txt，Mask 策略",
         "调整 keep_prefix=4, keep_suffix=2",
         "手机号脱敏为 1380*****00 形式",
         "sample.txt", "P2", "未覆盖", "", ""),

        # ===== 类型过滤 =====
        ("T01", "类型过滤", "全部类型启用",
         "导入 sample.csv",
         "默认状态（全部勾选）",
         "手机+身份证+邮箱+姓名均有高亮",
         "sample.csv", "P0", "未覆盖", "", ""),

        ("T02", "类型过滤", "只启用手机号",
         "导入 sample.csv",
         "取消勾选除 Phone 外的所有类型",
         "只有手机号列高亮，其他列无高亮",
         "sample.csv", "P1", "未覆盖", "", ""),

        ("T03", "类型过滤", "全关再全开",
         "导入 sample.csv",
         "Hide All → Show All",
         "高亮数恢复到 T01 相同值",
         "sample.csv", "P1", "未覆盖", "", ""),

        ("T04", "类型过滤", "关闭单一类型",
         "导入 sample.csv",
         "取消勾选 IdCard",
         "身份证列不高亮，其余不变",
         "sample.csv", "P2", "未覆盖", "", ""),

        # ===== 字典与白名单 =====
        ("D01", "字典/白名单", "添加字典条目",
         "导入 sample.txt（含'阿里巴巴'）",
         "右侧字典面板添加 '阿里巴巴' → OrgName",
         "1. 高亮数增加\n2. '阿里巴巴集团控股有限公司' 被高亮",
         "sample.txt", "P1", "未覆盖", "", ""),

        ("D02", "字典/白名单", "删除字典条目",
         "D01 完成后",
         "删除 '阿里巴巴' 条目",
         "'阿里巴巴' 不再高亮",
         "sample.txt", "P1", "未覆盖", "", ""),

        ("D03", "字典/白名单", "白名单排除",
         "导入 sample.txt，13800138000 已被高亮",
         "白名单添加 '13800138000'",
         "1. 该手机号不再高亮\n2. 其他手机号(13912345678)仍高亮\n3. 总高亮数减少",
         "sample.txt", "P1", "未覆盖", "", ""),

        ("D04", "字典/白名单", "字典+白名单组合",
         "导入 sample.txt",
         "1. 字典加 '某公司'→OrgName\n2. 白名单加 '张三'",
         "'某公司'新增高亮，'张三'消失",
         "sample.txt", "P2", "未覆盖", "", ""),

        ("D05", "字典/白名单", "模糊匹配",
         "导入 sample.txt",
         "字典添加 '阿里' (Fuzzy 模式)",
         "'阿里巴巴集团控股有限公司' 被匹配高亮",
         "sample.txt", "P2", "未覆盖", "", ""),

        # ===== 列级规则 =====
        ("L01", "列级规则", "自动列推断",
         "导入 sample.xlsx",
         "检查策略面板的列推断结果",
         "1. 手机号列识别为 Phone\n2. 身份证列识别为 IdCard\n3. 邮箱列识别为 Email",
         "sample.xlsx", "P1", "未覆盖", "", ""),

        ("L02", "列级规则", "修改列策略",
         "L01 完成后",
         "手机号列改为 Replace 策略",
         "该列脱敏结果从掩码变为假数据，其他列不变",
         "sample.xlsx", "P1", "未覆盖", "", ""),

        ("L03", "列级规则", "覆盖列类型为不敏感",
         "L01 完成后",
         "地址列改为不敏感",
         "地址列不再脱敏",
         "sample.xlsx", "P2", "未覆盖", "", ""),

        ("L04", "列级规则", "导入导出列规则",
         "L02 完成后",
         "导出规则 → 新工作区导入规则",
         "规则恢复，同一文件脱敏效果一致",
         "sample.xlsx", "P2", "未覆盖", "", ""),

        # ===== 一致性替换 =====
        ("K01", "一致性替换", "同文件一致性",
         "导入含重复手机号的 csv（Replace 策略）",
         "检查同一手机号出现多次的替换结果",
         "同一手机号 → 同一假数据（如 13800138000 在5行中替换结果相同）",
         "sample.csv", "P0", "未覆盖", "", "csv 中手机号每人不同，需创建含重复值的 fixture"),

        ("K02", "一致性替换", "跨文件一致性",
         "同一工作区",
         "1. 导入 file_a.csv (含13800138000)\n2. 导出\n3. 导入 file_b.csv (也含13800138000)",
         "两个文件中 13800138000 的替换结果相同",
         "需两个含相同手机号的文件", "P1", "未覆盖", "", "需创建专用 fixture"),

        ("K03", "一致性替换", "别名组一致性",
         "已创建工作区，导入含 '阿里巴巴' 和 '阿里' 的文件",
         "创建别名组：'阿里巴巴' + '阿里'",
         "两者替换为同一假名",
         "sample.txt", "P2", "未覆盖", "", ""),

        # ===== 还原 =====
        ("R01", "还原", "工作区还原",
         "已完成 Replace 脱敏并导出",
         "1. 回到 dropzone\n2. 点'从工作区还原'\n3. 选择脱敏后文件",
         "1. 进入还原视图\n2. 匹配数 > 0\n3. 还原后手机号恢复为 13800138000",
         "sample.txt", "P0", "部分覆盖", "test_restore.py",
         "当前只测按钮可见，未执行完整还原"),

        ("R02", "还原", "AI 回复还原",
         "已完成 Replace 脱敏",
         "1. 复制脱敏后文本\n2. 点'AI回复还原'\n3. 粘贴文本",
         "1. 假数据被替换回真实数据\n2. 匹配数 > 0",
         "sample.txt", "P0", "未覆盖", "", ""),

        ("R03", "还原", "Mask 不可逆验证",
         "已完成 Mask 脱敏",
         "尝试从工作区还原",
         "匹配数 = 0 或提示不可逆",
         "sample.txt", "P1", "未覆盖", "", ""),

        ("R04", "还原", "历史记录还原",
         "已完成多次脱敏",
         "从处理历史选择某条记录还原",
         "还原结果正确",
         "sample.txt", "P2", "未覆盖", "", ""),

        # ===== 批量处理 =====
        ("B01", "批量处理", "批量导入",
         "已创建工作区",
         "同时拖入 3 个文件",
         "FileQueue 显示 3 项，状态为 pending",
         "batch/*.{xlsx,csv,docx}", "P1", "未覆盖", "", ""),

        ("B02", "批量处理", "逐个导出",
         "B01 完成后",
         "对每个文件：等待处理 → 导出并下一个",
         "每个文件都生成导出文件，队列状态依次变为 confirmed",
         "batch/*", "P1", "未覆盖", "", ""),

        ("B03", "批量处理", "中途跳过",
         "B01 完成后",
         "处理第1个文件后点返回",
         "剩余文件保持 pending，回到 dropzone",
         "batch/*", "P2", "未覆盖", "", ""),

        ("B04", "批量处理", "混合格式",
         "已创建工作区",
         "同时导入 xlsx + csv + docx",
         "每种格式都正常处理和导出",
         "batch/*", "P2", "未覆盖", "", ""),

        # ===== 工作区管理 =====
        ("W01", "工作区管理", "创建工作区",
         "无",
         "Cmd+N",
         "工作区列表 +1",
         "", "P1", "部分覆盖", "test_workspace_crud.py",
         "当前已覆盖"),

        ("W02", "工作区管理", "重命名工作区",
         "已有工作区",
         "双击名称 → 输入'新名称' → Enter",
         "名称更新为'新名称'",
         "", "P1", "未覆盖", "", ""),

        ("W03", "工作区管理", "删除工作区",
         "已有 ≥ 2 个工作区",
         "删除当前工作区",
         "列表 -1，切换到其他工作区或 empty 视图",
         "", "P1", "未覆盖", "", ""),

        ("W04", "工作区管理", "切换工作区配置隔离",
         "工作区A (Mask) + 工作区B (Replace)",
         "在 A 和 B 之间切换",
         "策略面板跟随切换，不互相污染",
         "", "P1", "未覆盖", "", ""),

        ("W05", "工作区管理", "剪贴板工作区",
         "无",
         "创建剪贴板工作区 → 粘贴含敏感信息的文本",
         "进入处理流程，识别出手机号等敏感信息",
         "", "P2", "未覆盖", "", ""),

        # ===== UI 交互 =====
        ("U01", "UI交互", "侧栏折叠/展开",
         "默认状态",
         "Cmd+Shift+L / Cmd+Shift+R",
         "左/右面板隐藏/显示",
         "", "P2", "未覆盖", "", ""),

        ("U02", "UI交互", "返回导航",
         "已完成脱敏（comparison 视图）",
         "点击返回按钮",
         "回到 dropzone 视图",
         "", "P1", "已覆盖", "test_restore.py",
         "test_desensitize_then_back_to_dropzone"),

        ("U03", "UI交互", "点击高亮项弹窗",
         "已完成识别",
         "点击某个高亮项",
         "弹出 Popover，含'加字典/加白名单/改策略'选项",
         "", "P2", "未覆盖", "", ""),

        ("U04", "UI交互", "手动选文本标记",
         "已完成识别",
         "选中未识别文本 → 工具栏添加",
         "选中文本变为新的敏感高亮项",
         "", "P2", "未覆盖", "", ""),
    ]

    for row in cases:
        ws.append(row)

    style_header(ws, len(headers))
    style_rows(ws, 2, len(cases) + 1, len(headers), priority_col=8, coverage_col=9)

    # 列宽
    widths = [8, 12, 20, 18, 35, 40, 22, 8, 10, 28, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"
    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(cases) + 1}"


# ===== Sheet 2: Fixture 数据基线 =====
def create_baseline_sheet(wb):
    ws = wb.create_sheet("Fixture数据基线")

    headers = ["Fixture文件", "敏感值", "敏感类型", "出现次数", "备注"]
    ws.append(headers)

    # sample.txt 中的已知敏感数据
    baseline = [
        # sample.txt
        ("sample.txt", "张三", "PersonName", 1, "NER 识别"),
        ("sample.txt", "13800138000", "Phone", 1, "正则识别"),
        ("sample.txt", "110101199001011234", "IdCard", 1, "正则识别"),
        ("sample.txt", "zhangsan@example.com", "Email", 1, "正则识别"),
        ("sample.txt", "北京市朝阳区建国路88号", "Address", 1, "NER 识别"),
        ("sample.txt", "李四", "PersonName", 1, "NER 识别"),
        ("sample.txt", "13912345678", "Phone", 1, "正则识别"),
        ("sample.txt", "320102198512152345", "IdCard", 1, "正则识别"),
        ("sample.txt", "lisi@test.com", "Email", 1, "正则识别"),
        ("sample.txt", "上海市浦东新区陆家嘴环路1000号", "Address", 1, "NER 识别"),
        ("sample.txt", "阿里巴巴集团控股有限公司", "OrgName", 1, "NER 识别"),
        ("sample.txt", "91330100799655058B", "CreditCode", 1, "正则识别"),

        # sample.csv（5行数据 × 5列）
        ("sample.csv", "张三", "PersonName", 1, "NER（表头'姓名'列）"),
        ("sample.csv", "李四", "PersonName", 1, ""),
        ("sample.csv", "王五", "PersonName", 1, ""),
        ("sample.csv", "赵六", "PersonName", 1, ""),
        ("sample.csv", "钱七", "PersonName", 1, ""),
        ("sample.csv", "13800138000", "Phone", 1, "正则"),
        ("sample.csv", "13912345678", "Phone", 1, ""),
        ("sample.csv", "15011112222", "Phone", 1, ""),
        ("sample.csv", "18688889999", "Phone", 1, ""),
        ("sample.csv", "17700001111", "Phone", 1, ""),
        ("sample.csv", "110101199001011234", "IdCard", 1, "正则"),
        ("sample.csv", "320102198512152345", "IdCard", 1, ""),
        ("sample.csv", "440106197803203456", "IdCard", 1, ""),
        ("sample.csv", "510105199207074567", "IdCard", 1, ""),
        ("sample.csv", "330102198911115678", "IdCard", 1, ""),
        ("sample.csv", "zhangsan@example.com", "Email", 1, "正则"),
        ("sample.csv", "lisi@test.com", "Email", 1, ""),
        ("sample.csv", "wangwu@demo.org", "Email", 1, ""),
        ("sample.csv", "zhaoliu@mail.cn", "Email", 1, ""),
        ("sample.csv", "qianqi@corp.io", "Email", 1, ""),
        ("sample.csv", "北京市朝阳区建国路88号", "Address", 1, "NER"),
        ("sample.csv", "上海市浦东新区陆家嘴环路1000号", "Address", 1, ""),
        ("sample.csv", "广州市天河区体育西路191号", "Address", 1, ""),
        ("sample.csv", "成都市武侯区人民南路四段1号", "Address", 1, ""),
        ("sample.csv", "杭州市西湖区文三路269号", "Address", 1, ""),

        # sample.xlsx（与 csv 数据相同）
        ("sample.xlsx", "(与 sample.csv 相同)", "", 0, "5行×5列，数据一致"),

        # sample.docx
        ("sample.docx", "张三", "PersonName", 1, "NER 识别"),
        ("sample.docx", "13800138000", "Phone", 1, "正则"),
        ("sample.docx", "110101199001011234", "IdCard", 1, "正则"),
        ("sample.docx", "zhangsan@example.com", "Email", 1, "正则"),
        ("sample.docx", "北京市朝阳区建国路88号", "Address", 1, "NER"),
        ("sample.docx", "李四", "PersonName", 1, "NER"),
        ("sample.docx", "13912345678", "Phone", 1, "正则"),
        ("sample.docx", "320102198512152345", "IdCard", 1, "正则"),
        ("sample.docx", "lisi@test.com", "Email", 1, "正则"),
        ("sample.docx", "上海市浦东新区陆家嘴环路1000号", "Address", 1, "NER"),
        ("sample.docx", "阿里巴巴集团控股有限公司", "OrgName", 1, "NER"),
    ]

    for row in baseline:
        ws.append(row)

    style_header(ws, len(headers))
    style_rows(ws, 2, len(baseline) + 1, len(headers))

    widths = [15, 35, 15, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(baseline) + 1}"


# ===== Sheet 3: 覆盖率统计 =====
def create_coverage_sheet(wb):
    ws = wb.create_sheet("覆盖率统计")

    headers = ["分类", "总用例", "已覆盖", "部分覆盖", "未覆盖", "覆盖率"]
    ws.append(headers)

    stats = [
        ("核心管道", 9, 0, 4, 5, "4/9 (部分)"),
        ("策略切换", 6, 0, 0, 6, "0/6"),
        ("类型过滤", 4, 0, 0, 4, "0/4"),
        ("字典/白名单", 5, 0, 0, 5, "0/5"),
        ("列级规则", 4, 0, 0, 4, "0/4"),
        ("一致性替换", 3, 0, 0, 3, "0/3"),
        ("还原", 4, 0, 1, 3, "1/4 (部分)"),
        ("批量处理", 4, 0, 0, 4, "0/4"),
        ("工作区管理", 5, 0, 1, 4, "1/5 (部分)"),
        ("UI交互", 4, 1, 0, 3, "1/4"),
        ("", "", "", "", "", ""),
        ("合计", 48, 1, 6, 41, "7/48 (14.6%)"),
    ]

    for row in stats:
        ws.append(row)

    style_header(ws, len(headers))
    style_rows(ws, 2, len(stats) + 1, len(headers))

    widths = [15, 10, 10, 10, 10, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ===== 主函数 =====
def main():
    wb = Workbook()
    create_testcase_sheet(wb)
    create_baseline_sheet(wb)
    create_coverage_sheet(wb)
    wb.save(OUTPUT)
    print(f"测试用例 Excel 已生成: {OUTPUT}")


if __name__ == "__main__":
    main()
