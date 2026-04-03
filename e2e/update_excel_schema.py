#!/usr/bin/env python3
"""一次性脚本：为 testcases.xlsx 新增 Sheet1 执行结果列 + Sheet2 断言模式列"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

EXCEL_PATH = Path(__file__).parent / "testcases.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

wb = load_workbook(str(EXCEL_PATH))

# === Sheet1: 新增 4 列（执行结果、失败原因、执行时间、截图路径）===
ws1 = wb["测试用例"]
new_headers = {12: "执行结果", 13: "失败原因", 14: "执行时间", 15: "截图路径"}
for col, header in new_headers.items():
    cell = ws1.cell(1, col, header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

# 给已有数据行填默认值
for row in range(2, ws1.max_row + 1):
    if ws1.cell(row, 1).value:
        ws1.cell(row, 12, "未执行")
        for col in range(12, 16):
            ws1.cell(row, col).border = THIN_BORDER
            ws1.cell(row, col).alignment = Alignment(wrap_text=True, vertical="top")

# 更新自动筛选范围
ws1.auto_filter.ref = f"A1:O{ws1.max_row}"

# === Sheet2: 新增"断言模式"列 ===
ws2 = wb["Fixture数据基线"]
cell = ws2.cell(1, 6, "断言模式")
cell.font = HEADER_FONT
cell.fill = HEADER_FILL
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.border = THIN_BORDER

# 根据备注自动填充断言模式
for row in range(2, ws2.max_row + 1):
    note = ws2.cell(row, 5).value or ""
    mode = "soft" if "NER" in note else "hard"
    ws2.cell(row, 6, mode)
    ws2.cell(row, 6).border = THIN_BORDER
    ws2.cell(row, 6).alignment = Alignment(vertical="top")

# 更新自动筛选范围
ws2.auto_filter.ref = f"A1:F{ws2.max_row}"

wb.save(str(EXCEL_PATH))
print(f"Excel schema 已更新: {EXCEL_PATH}")
