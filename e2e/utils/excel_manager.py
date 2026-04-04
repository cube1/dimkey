"""Dimkey E2E 测试用例 Excel 管理器

读写 testcases.xlsx：
- Sheet1 "测试用例": 用例定义和执行结果
- Sheet2 "Fixture数据基线": fixture 文件的期望敏感值
"""

from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment

EXCEL_PATH = Path(__file__).resolve().parent.parent / "testcases.xlsx"

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")

# Sheet1 列索引（1-based）
COL_ID = 1
COL_CATEGORY = 2
COL_SCENARIO = 3
COL_PRECONDITION = 4
COL_STEPS = 5
COL_EXPECTED = 6
COL_FIXTURE = 7
COL_PRIORITY = 8
COL_COVERAGE = 9
COL_TEST_FILE = 10
COL_NOTE = 11
COL_EXEC_RESULT = 12
COL_FAIL_REASON = 13
COL_EXEC_TIME = 14
COL_SCREENSHOT = 15

# Sheet2 列索引（1-based）
BL_COL_FIXTURE = 1
BL_COL_VALUE = 2
BL_COL_TYPE = 3
BL_COL_COUNT = 4
BL_COL_NOTE = 5
BL_COL_ASSERT_MODE = 6


def _load_wb():
    """加载工作簿"""
    return load_workbook(str(EXCEL_PATH))


def _save_wb(wb):
    """保存工作簿"""
    wb.save(str(EXCEL_PATH))


def _style_row(ws, row, col_count):
    """给新行添加边框和换行"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = THIN_BORDER
        cell.alignment = WRAP


def read_testcases() -> list[dict]:
    """读取 Sheet1 所有用例"""
    wb = _load_wb()
    ws = wb["测试用例"]
    cases = []
    for row in range(2, ws.max_row + 1):
        case_id = ws.cell(row, COL_ID).value
        if not case_id:
            continue
        cases.append({
            "id": case_id,
            "category": ws.cell(row, COL_CATEGORY).value,
            "scenario": ws.cell(row, COL_SCENARIO).value,
            "precondition": ws.cell(row, COL_PRECONDITION).value,
            "steps": ws.cell(row, COL_STEPS).value,
            "expected": ws.cell(row, COL_EXPECTED).value,
            "fixture": ws.cell(row, COL_FIXTURE).value,
            "priority": ws.cell(row, COL_PRIORITY).value,
            "coverage": ws.cell(row, COL_COVERAGE).value,
            "test_file": ws.cell(row, COL_TEST_FILE).value,
            "note": ws.cell(row, COL_NOTE).value,
            "exec_result": ws.cell(row, COL_EXEC_RESULT).value,
            "fail_reason": ws.cell(row, COL_FAIL_REASON).value,
            "exec_time": ws.cell(row, COL_EXEC_TIME).value,
            "screenshot": ws.cell(row, COL_SCREENSHOT).value,
            "row": row,
        })
    wb.close()
    return cases


def _next_id(ws, prefix: str) -> str:
    """根据前缀计算下一个用例 ID，如 C10, S07"""
    max_num = 0
    for row in range(2, ws.max_row + 1):
        cell_id = ws.cell(row, COL_ID).value
        if cell_id and cell_id.startswith(prefix):
            try:
                num = int(cell_id[len(prefix):])
                max_num = max(max_num, num)
            except ValueError:
                pass
    return f"{prefix}{max_num + 1:02d}"


# 分类 → ID 前缀映射
CATEGORY_PREFIX = {
    "核心管道": "C",
    "策略切换": "S",
    "类型过滤": "T",
    "字典/白名单": "D",
    "列级规则": "L",
    "一致性替换": "K",
    "还原": "R",
    "批量处理": "B",
    "工作区管理": "W",
    "UI交互": "U",
}


def add_testcase(case: dict) -> str:
    """写入新用例行，返回分配的用例 ID

    case keys: category, scenario, precondition, steps, expected,
               fixture, priority, test_file, note
    """
    wb = _load_wb()
    ws = wb["测试用例"]

    prefix = CATEGORY_PREFIX.get(case.get("category", ""), "X")
    case_id = _next_id(ws, prefix)

    row = ws.max_row + 1
    ws.cell(row, COL_ID, case_id)
    ws.cell(row, COL_CATEGORY, case.get("category", ""))
    ws.cell(row, COL_SCENARIO, case.get("scenario", ""))
    ws.cell(row, COL_PRECONDITION, case.get("precondition", ""))
    ws.cell(row, COL_STEPS, case.get("steps", ""))
    ws.cell(row, COL_EXPECTED, case.get("expected", ""))
    ws.cell(row, COL_FIXTURE, case.get("fixture", ""))
    ws.cell(row, COL_PRIORITY, case.get("priority", "P1"))
    ws.cell(row, COL_COVERAGE, "未覆盖")
    ws.cell(row, COL_TEST_FILE, case.get("test_file", ""))
    ws.cell(row, COL_NOTE, case.get("note", ""))
    ws.cell(row, COL_EXEC_RESULT, "未执行")

    _style_row(ws, row, 15)
    _save_wb(wb)
    return case_id


def read_baseline(fixture_file: str) -> list[dict]:
    """读取 Sheet2 中某 fixture 的所有期望敏感值

    返回: [{"value": "13800138000", "type": "Phone", "assert_mode": "hard"}, ...]
    """
    wb = _load_wb()
    ws = wb["Fixture数据基线"]
    items = []
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, BL_COL_FIXTURE).value == fixture_file:
            value = ws.cell(row, BL_COL_VALUE).value
            if not value or str(value).startswith("("):
                continue
            note = ws.cell(row, BL_COL_NOTE).value or ""
            assert_mode_cell = ws.cell(row, BL_COL_ASSERT_MODE).value
            if assert_mode_cell:
                mode = assert_mode_cell
            elif "NER" in note:
                mode = "soft"
            else:
                mode = "hard"
            items.append({
                "value": str(value).strip().replace('\xa0', ' '),
                "type": ws.cell(row, BL_COL_TYPE).value or "",
                "assert_mode": mode,
            })
    wb.close()
    return items


def add_baseline(fixture_file: str, items: list[dict]):
    """写入新基线条目

    items: [{"value": "xxx", "type": "Phone", "count": 1, "note": "正则", "assert_mode": "hard"}, ...]
    """
    wb = _load_wb()
    ws = wb["Fixture数据基线"]
    for item in items:
        row = ws.max_row + 1
        ws.cell(row, BL_COL_FIXTURE, fixture_file)
        ws.cell(row, BL_COL_VALUE, item["value"])
        ws.cell(row, BL_COL_TYPE, item.get("type", ""))
        ws.cell(row, BL_COL_COUNT, item.get("count", 1))
        ws.cell(row, BL_COL_NOTE, item.get("note", ""))
        ws.cell(row, BL_COL_ASSERT_MODE, item.get("assert_mode", "hard"))
        _style_row(ws, row, 6)
    _save_wb(wb)


def update_result(case_id: str, result: dict):
    """回写执行结果到 Sheet1

    result keys: exec_result ("通过"/"失败"), fail_reason, screenshot, coverage, test_file
    """
    wb = _load_wb()
    ws = wb["测试用例"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, COL_ID).value == case_id:
            if result.get("exec_result") is not None:
                ws.cell(row, COL_EXEC_RESULT, result.get("exec_result", ""))
            if result.get("fail_reason") is not None:
                ws.cell(row, COL_FAIL_REASON, result.get("fail_reason", ""))
            if result.get("exec_result") is not None or result.get("fail_reason") is not None:
                ws.cell(row, COL_EXEC_TIME, datetime.now().strftime("%Y-%m-%d %H:%M"))
            if result.get("screenshot"):
                ws.cell(row, COL_SCREENSHOT, result["screenshot"])
            if result.get("test_file"):
                ws.cell(row, COL_TEST_FILE, result["test_file"])
            if result.get("coverage"):
                ws.cell(row, COL_COVERAGE, result["coverage"])
                fill_map = {
                    "已覆盖": PatternFill(start_color="D3F9D8", end_color="D3F9D8", fill_type="solid"),
                    "部分覆盖": PatternFill(start_color="FFF3BF", end_color="FFF3BF", fill_type="solid"),
                    "未覆盖": PatternFill(start_color="FFE3E3", end_color="FFE3E3", fill_type="solid"),
                }
                if result["coverage"] in fill_map:
                    ws.cell(row, COL_COVERAGE).fill = fill_map[result["coverage"]]
            if result.get("exec_result") == "通过":
                ws.cell(row, COL_EXEC_RESULT).fill = PatternFill(
                    start_color="D3F9D8", end_color="D3F9D8", fill_type="solid")
            elif result.get("exec_result") == "失败":
                ws.cell(row, COL_EXEC_RESULT).fill = PatternFill(
                    start_color="FFE3E3", end_color="FFE3E3", fill_type="solid")
            break
    _save_wb(wb)
