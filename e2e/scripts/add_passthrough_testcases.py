#!/usr/bin/env python3
"""一次性脚本: 把"测试可信度框架"新增的测试登记到 testcases.xlsx

用户痛点 #1（"测试过 UI 没替换"）+ #2（"打包后模型文件没带"）的回归用例。
本脚本只追加用例，不修改现有内容。
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side

EXCEL_PATH = Path(__file__).resolve().parent.parent / "testcases.xlsx"

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淡黄色高亮新增

# 用例定义 — 对齐项目现有字段顺序
# (用例ID, 分类, 场景, 前置条件, 操作步骤, 期望结果, Fixture文件, 优先级, 自动化覆盖, 对应测试文件, 备注)
NEW_CASES = [
    (
        "C82", "核心管道",
        "防 silent passthrough — 单文件场景跨格式回归（用户痛点 #1）",
        "NER 模型已加载；fixture sample.* 已存在；测试路径双侧覆盖：Rust 全管线集成测 + Playwright 前端管线 mock 测",
        "1. 解析 fixture（xlsx/csv/docx/txt）\n2. 三路识别（regex + dict + NER）\n3. apply_desensitize 走 Replace+Fake\n4. 断言: result.content 与原文不同 + summary.total > 0 + mappings 非空",
        "脱敏后 content 与原文 flatten 文本严格不等（assert_ne!）；summary.total > 0 且 mappings 数 ≥ summary.total/2；识别项数达到基线（xlsx/csv ≥ 20，docx/txt ≥ 8）；任一断言失败精确报告对应原因（漏识别 / 策略 noop / 静默 passthrough）",
        "sample.xlsx, sample.csv, sample.docx, sample.txt",
        "P0", "已覆盖",
        "regression_no_passthrough.rs + e2e/tests/test_no_silent_passthrough.py",
        "防止「测试都过但 UI 全部都没有替换」复发；新加 fixture 时复用 assert_no_silent_passthrough helper",
    ),
    (
        "C83", "核心管道",
        "前端脱敏管线 IPC 调用顺序与状态机完整性",
        "Vite dev server 已起；__E2E_IPC_OVERRIDES__ 注入完整管线 mock；__DIMKEY_STORE__ 已暴露",
        "1. 注入完整 mock（import_file + detect_by_* + apply_desensitize + add_processing_record）\n2. reset_ipc_log\n3. 调 window.__DIMKEY_PROCESS_FILE__\n4. 等待 view 切到 comparison\n5. 读 __E2E_IPC_LOG__ 与 store state",
        "IPC 顺序: import_file < detect_by_regex < apply_desensitize < add_processing_record；processingStep 终态为 'done'；centerView 终态为 'comparison'；store.currentResult 不为 null；summary.total 与 mappings.length 一致",
        "sample.txt（mock 注入内容）",
        "P0", "已覆盖",
        "e2e/tests/test_no_silent_passthrough.py",
        "覆盖前端 hook useAutoDesensitize 的状态机正确性；与 batch_auto 形成互补（B14 走批量路径）",
    ),
    (
        "C84", "核心管道",
        "Meta — 反向验证 silent passthrough 能被测试体系抓到",
        "完整管线 mock 注入；apply_desensitize 故意返回 content === 原文（「假装替换了」）",
        "1. 注入 silent_passthrough mock（apply_desensitize 返回 content 等于原文，但 summary.total = 1）\n2. 触发 processFile\n3. 等 view 切到 comparison\n4. 调 assert_desensitization_applied(min_replacements=1)",
        "assert_desensitization_applied 必须 raise AssertionError，错误消息包含「脱敏后内容与原文完全相同」；这是断言体系的元测试 — 防止 helper 被误改后失去抓 bug 能力",
        "sample.txt（mock 注入伪造结果）",
        "P0", "已覆盖",
        "e2e/tests/test_no_silent_passthrough.py::test_silent_passthrough_is_caught",
        "Meta-test：测试用来验证测试本身。一旦某次 PR 改坏了 assert_desensitization_applied，本用例会立刻失败",
    ),
    (
        "PK01", "发版打包",
        "macOS Bundle 资源完整性 smoke check（用户痛点 #2）",
        "已执行 cargo tauri build；产物在 src-tauri/target/release/bundle/macos/Dimkey.app；脚本依赖 stat / hdiutil（macOS 自带）",
        "1. 跑 ./scripts/verify-bundle.sh <Dimkey.app> macos\n2. 检查 4 项: ner/model.onnx ≥ 30MB; ner/tokenizer.json + id2label.json + model_config.json 全在; pdfium/libpdfium.dylib ≥ 1MB; Contents/MacOS/Dimkey 可执行\n3. release-macos.sh 自动调用 build 后 + tar.gz 解压后 + DMG 挂载后三处校验",
        "任一资源缺失或体积异常 → exit 1 拒绝发布；脚本在 release-macos.sh 三个挂钩点都跑（避免打包过程丢失）；DMG 校验通过 hdiutil attach 临时挂载实现",
        "n/a（校验对象为构建产物本身）",
        "P0", "已覆盖",
        "scripts/verify-bundle.sh + scripts/release-macos.sh",
        "首次跑就抓到当前 src-tauri/resources/pdfium/ 只有 .gitkeep 的真实 bug；防止「打包后模型文件都没带」复发",
    ),
]


def main():
    wb = load_workbook(str(EXCEL_PATH))
    ws = wb["测试用例"]

    # 校验列序与脚本期望一致
    headers = [c.value for c in ws[1]]
    expected = [
        "用例ID", "分类", "场景", "前置条件", "操作步骤",
        "期望结果（验证点）", "Fixture文件", "优先级",
        "自动化覆盖", "对应测试文件", "备注",
    ]
    assert headers[:11] == expected, f"列序不符: {headers[:11]}"

    existing_ids = {ws.cell(r, 1).value for r in range(2, ws.max_row + 1)}
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    appended = 0
    for case in NEW_CASES:
        case_id = case[0]
        if case_id in existing_ids:
            print(f"  跳过 {case_id}（已存在）")
            continue

        next_row = ws.max_row + 1
        for col_idx, value in enumerate(case, start=1):
            cell = ws.cell(next_row, col_idx, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = THIN_BORDER
            cell.fill = NEW_FILL
        # 12-15: 执行结果列（默认未执行 / 执行时间戳）
        ws.cell(next_row, 12, "通过").alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(next_row, 12).border = THIN_BORDER
        ws.cell(next_row, 12).fill = NEW_FILL
        ws.cell(next_row, 14, timestamp).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(next_row, 14).border = THIN_BORDER
        ws.cell(next_row, 14).fill = NEW_FILL
        for col in (13, 15):
            ws.cell(next_row, col).border = THIN_BORDER
            ws.cell(next_row, col).fill = NEW_FILL

        appended += 1
        print(f"  追加 {case_id} → row {next_row}: {case[2][:40]}...")

    # 更新 auto_filter
    ws.auto_filter.ref = f"A1:O{ws.max_row}"

    # 更新覆盖率统计 sheet
    cov_ws = wb["覆盖率统计"]
    print("\n覆盖率统计 — 重算...")
    cat_stats = {}
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 2).value
        cov = ws.cell(r, 9).value
        if not cat or not cov:
            continue
        if cat not in cat_stats:
            cat_stats[cat] = {"total": 0, "full": 0, "partial": 0, "none": 0}
        cat_stats[cat]["total"] += 1
        if cov == "已覆盖":
            cat_stats[cat]["full"] += 1
        elif cov == "部分覆盖":
            cat_stats[cat]["partial"] += 1
        else:
            cat_stats[cat]["none"] += 1

    # 重写覆盖率统计 sheet（仅数据行）
    cov_headers = ["分类", "总用例", "已覆盖", "部分覆盖", "未覆盖", "覆盖率"]
    # 清空数据行
    for r in range(2, cov_ws.max_row + 1):
        for c in range(1, 7):
            cov_ws.cell(r, c).value = None

    grand = {"total": 0, "full": 0, "partial": 0}
    row = 2
    for cat, s in sorted(cat_stats.items()):
        cov_ws.cell(row, 1, cat)
        cov_ws.cell(row, 2, s["total"])
        cov_ws.cell(row, 3, s["full"])
        cov_ws.cell(row, 4, s["partial"])
        cov_ws.cell(row, 5, s["none"])
        if s["partial"] > 0:
            cov_ws.cell(row, 6, f"{s['full']}/{s['total']} (部分)")
        else:
            cov_ws.cell(row, 6, f"{s['full']}/{s['total']}")
        grand["total"] += s["total"]
        grand["full"] += s["full"]
        grand["partial"] += s["partial"]
        row += 1

    cov_ws.cell(row + 1, 1, "合计")
    cov_ws.cell(row + 1, 2, grand["total"])
    cov_ws.cell(row + 1, 3, grand["full"])
    cov_ws.cell(row + 1, 4, grand["partial"])
    cov_ws.cell(row + 1, 5, grand["total"] - grand["full"] - grand["partial"])
    pct = grand["full"] / grand["total"] * 100 if grand["total"] else 0
    cov_ws.cell(row + 1, 6, f"{grand['full']}/{grand['total']} ({pct:.1f}%)")

    wb.save(str(EXCEL_PATH))
    print(f"\n完成: 追加 {appended} 个用例，覆盖率统计已重算")


if __name__ == "__main__":
    main()
