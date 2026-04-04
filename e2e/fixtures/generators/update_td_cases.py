#!/usr/bin/env python3
"""更新 T01-T04 和 D01-D05 用例的详细描述、步骤、期望结果、fixture"""
import sys
sys.path.insert(0, "e2e")

from openpyxl import load_workbook
from pathlib import Path

EXCEL = Path("e2e/testcases.xlsx")
wb = load_workbook(str(EXCEL))
ws = wb["测试用例"]

# 用例 ID → 行号 映射
id_to_row = {}
for row in range(2, ws.max_row + 1):
    cid = ws.cell(row, 1).value
    if cid:
        id_to_row[cid] = row

# 列索引
COL_SCENARIO = 3
COL_PRECONDITION = 4
COL_STEPS = 5
COL_EXPECTED = 6
COL_FIXTURE = 7

updates = {
    "T01": {
        "scenario": "全类型启用: 识别结果包含所有敏感类型",
        "precondition": "fixture: scenarios/csv/员工信息表.csv（含 Phone/IdCard/Email/BankCard/Address/OrgName）",
        "steps": "1. 导入员工信息表.csv\n2. enabled_types 传入全部类型\n3. 调用 detect_by_regex\n4. 检查结果",
        "expected": "识别结果中 Phone >= 8, IdCard >= 8, Email >= 8, BankCard >= 8 均存在",
        "fixture": "scenarios/csv/员工信息表.csv",
    },
    "T02": {
        "scenario": "只启用手机号: 识别结果仅含 Phone 类型",
        "precondition": "fixture: scenarios/csv/员工信息表.csv",
        "steps": '1. 导入员工信息表.csv\n2. enabled_types = ["Phone"]\n3. 调用 detect_by_regex\n4. 检查结果类型',
        "expected": "识别结果中所有 item 的 sensitive_type 均为 Phone；不包含 IdCard/Email/BankCard 等其他类型",
        "fixture": "scenarios/csv/员工信息表.csv",
    },
    "T03": {
        "scenario": "全关再全开: 关闭所有类型后识别为空，重新全开后恢复",
        "precondition": "fixture: scenarios/csv/员工信息表.csv",
        "steps": "1. 导入员工信息表.csv\n2. enabled_types = []（空列表）\n3. 调用 detect_by_regex → 验证结果为空\n4. enabled_types 恢复为全部类型\n5. 再次调用 detect_by_regex → 验证结果非空",
        "expected": "第一次识别结果为空列表；第二次识别结果包含 Phone/IdCard/Email 等多种类型",
        "fixture": "scenarios/csv/员工信息表.csv",
    },
    "T04": {
        "scenario": "关闭单一类型: 仅 IdCard 被排除，其他类型不受影响",
        "precondition": "fixture: scenarios/csv/员工信息表.csv",
        "steps": '1. 导入员工信息表.csv\n2. enabled_types = 全部类型但去掉 "IdCard"\n3. 调用 detect_by_regex\n4. 检查结果',
        "expected": "识别结果中不包含 IdCard 类型；Phone/Email/BankCard 等数量与全量识别一致",
        "fixture": "scenarios/csv/员工信息表.csv",
    },
    "D01": {
        "scenario": "添加字典条目后命中: 新增自定义词条应在识别结果中出现",
        "precondition": "fixture: sample.txt（含常规敏感数据）",
        "steps": '1. 构造 DictEntry: text="阿里巴巴集团控股有限公司", type=Custom("公司名"), mode=Exact\n2. 调用 detect_by_dict(content, [entry])\n3. 检查结果',
        "expected": "识别结果中包含 text='阿里巴巴集团控股有限公司'、source=Dict 的 SensitiveItem",
        "fixture": "sample.txt",
    },
    "D02": {
        "scenario": "删除字典条目后不再命中: 移除词条后该文本不被字典引擎识别",
        "precondition": "fixture: sample.txt",
        "steps": "1. 调用 detect_by_dict(content, [entry]) → 验证命中\n2. 调用 detect_by_dict(content, []) → 空词典\n3. 检查结果",
        "expected": "第一次有命中结果；第二次（空词典）识别结果中无任何 Dict source 的 item",
        "fixture": "sample.txt",
    },
    "D03": {
        "scenario": "白名单排除: 将已识别的敏感值加入白名单后，该值不再出现在最终结果中",
        "precondition": "fixture: sample.txt（含手机号 13800138000）",
        "steps": '1. 全量识别 sample.txt → 确认包含 "13800138000"\n2. 将 "13800138000" 加入白名单\n3. 过滤后检查结果',
        "expected": '过滤后识别结果中不包含 text="13800138000" 的 item；其他敏感值不受影响',
        "fixture": "sample.txt",
    },
    "D04": {
        "scenario": "字典+白名单组合: 字典命中的条目若在白名单中则被排除",
        "precondition": "fixture: sample.txt",
        "steps": '1. 添加字典条目 "阿里巴巴集团控股有限公司"\n2. 同时将 "阿里巴巴集团控股有限公司" 加入白名单\n3. 合并识别结果并应用白名单过滤',
        "expected": "字典引擎命中该词条，但白名单过滤后最终结果中不包含该词条",
        "fixture": "sample.txt",
    },
    "D05": {
        "scenario": "模糊匹配: 字典条目设置 Fuzzy 模式后忽略大小写匹配",
        "precondition": "fixture: sample.txt（含 zhangsan@example.com）",
        "steps": '1. 构造 DictEntry: text="ZhangSan@Example.COM", mode=Fuzzy\n2. 调用 detect_by_dict\n3. 检查结果',
        "expected": '模糊匹配命中 "zhangsan@example.com"（忽略大小写）',
        "fixture": "sample.txt",
    },
}

for case_id, fields in updates.items():
    row = id_to_row.get(case_id)
    if not row:
        print(f"  ⚠ {case_id} 未找到")
        continue
    ws.cell(row, COL_SCENARIO, fields["scenario"])
    ws.cell(row, COL_PRECONDITION, fields["precondition"])
    ws.cell(row, COL_STEPS, fields["steps"])
    ws.cell(row, COL_EXPECTED, fields["expected"])
    ws.cell(row, COL_FIXTURE, fields["fixture"])
    print(f"  ✓ {case_id}: {fields['scenario']}")

wb.save(str(EXCEL))
print(f"\n已更新 {len(updates)} 条用例")
