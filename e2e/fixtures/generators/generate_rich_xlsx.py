#!/usr/bin/env python3
"""生成多行业场景的丰富 Excel 测试文件"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def styled_header(ws, headers, fill_color="4472C4"):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border


def create_case_register():
    """律所案件登记表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "案件登记"

    headers = ["案件编号", "案件类型", "委托人姓名", "委托人身份证", "委托人电话",
               "委托人邮箱", "委托人地址", "对方当事人", "对方联系方式",
               "主办律师", "律师电话", "案件标的(万元)", "委托日期", "状态"]
    styled_header(ws, headers)

    data = [
        ["AL-2024-0301", "离婚纠纷", "张丽芳", "110101199005123456", "13801234567",
         "zhanglifang@qq.com", "北京市朝阳区望京SOHO T1-2301",
         "陈明辉", "13901234568", "林志远", "13607012345", "120", "2024-03-01", "审理中"],
        ["AL-2024-0302", "合同纠纷", "苏州星河智能制造科技有限公司", "91320500MA1WXYZ123", "13812345678",
         "gumingyuan@xhzn.com", "苏州市工业园区星湖街328号创意产业园6号楼",
         "无锡创达机械有限公司", "0510-85123456", "陈国栋", "13512345678", "350", "2024-03-05", "调解中"],
        ["AL-2024-0315", "劳动争议", "方志诚", "360102199507151234", "15879012345",
         "fangzhicheng@outlook.com", "南昌市东湖区八一大道369号万达星城B区7栋1602",
         "南昌红谷滩创新科技有限公司", "0791-88123456", "林志远", "13607012345", "16.8", "2024-03-15", "仲裁中"],
        ["AL-2024-0401", "交通事故", "梁美玲", "440106199103151234", "15823456789",
         "liangmeiling@163.com", "广州市海珠区新港中路356号金逸华庭3栋1801",
         "谢国强", "13823456789", "何思雨", "13723456789", "22.55", "2024-04-01", "理赔中"],
        ["AL-2024-0412", "房屋买卖", "刘晓燕", "430104198805123456", "13974561234",
         "liuxiaoyan@163.com", "深圳市南山区蛇口街道海上世界双玺花园3栋2801",
         "深圳市锦程房地产开发有限公司", "0755-86123456", "陈国栋", "13512345678", "685", "2024-04-12", "起诉中"],
        ["AL-2024-0503", "知识产权", "杭州盛达贸易有限公司", "91330100MA2BXYZ789", "13867012345",
         "jiangwenbin72@sina.com", "杭州市江干区钱江新城来福士广场T2-2501",
         "杭州仿冒电子商务有限公司", "15867012345", "赵启明", "13823456789", "200", "2024-05-03", "取证中"],
        ["AL-2024-0515", "民间借贷", "唐俊杰", "110105199208151234", "18610012345",
         "tangjunjie@gmail.com", "北京市海淀区知春路甲48号盈都大厦C座601",
         "范国庆", "13901234567", "何思雨", "13723456789", "80", "2024-05-15", "执行中"],
        ["AL-2024-0601", "公司股权", "顾明远", "320505198803121234", "13812345678",
         "gumingyuan@xhzn.com", "苏州市工业园区星湖街218号湖畔花园12栋601",
         "沈志豪", "13912345678", "林志远", "13607012345", "500", "2024-06-01", "协商中"],
    ]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [16, 12, 22, 22, 15, 28, 40, 24, 18, 10, 15, 14, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/律所案件登记表.xlsx")
    print("已生成: 律所案件登记表.xlsx")


def create_patient_register():
    """医院患者登记"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "患者登记"

    headers = ["病历号", "姓名", "性别", "年龄", "身份证号", "联系电话",
               "家庭住址", "工作单位", "紧急联系人", "紧急联系电话",
               "就诊科室", "主治医师", "初步诊断", "医保卡号"]
    styled_header(ws, headers, "2E7D32")

    data = [
        ["MR-2024-0001", "蒋文斌", "男", 52, "330102197206153456", "13867012345",
         "杭州市上城区庆春路218号嘉德广场6幢1203", "杭州盛达贸易有限公司",
         "王婷", "13757012345", "心血管内科", "周明华", "不稳定型心绞痛", "ZJ330102197206150019"],
        ["MR-2024-0002", "黄雅琴", "女", 34, "310101199008151234", "13512345670",
         "上海市徐汇区衡山路880号永嘉庭5号602", "上海博雅教育咨询有限公司",
         "黄父", "13612345670", "妇产科", "陈丽华", "先兆流产", "SH310101199008150018"],
        ["MR-2024-0003", "程建军", "男", 45, "420102197907231234", "13607012346",
         "武汉市武昌区东湖路100号碧湖花园8栋301", "武汉长江通信技术有限公司",
         "程妻 李艳", "13807012346", "骨科", "张明德", "腰椎间盘突出", "HB420102197907230015"],
        ["MR-2024-0004", "沈雨薇", "女", 28, "320102199603081234", "15212345678",
         "南京市鼓楼区中山北路405号恒基中心21楼", "南京网易有限公司",
         "沈母", "15312345678", "皮肤科", "林晓峰", "慢性荨麻疹", "JS320102199603080016"],
        ["MR-2024-0005", "郭志伟", "男", 63, "440102196108051234", "13323456789",
         "广州市越秀区环市东路339号广东国际大厦B座", "退休",
         "郭子 郭磊", "18623456789", "神经内科", "王芳", "帕金森病", "GD440102196108050012"],
        ["MR-2024-0006", "韩雪梅", "女", 41, "510102198303121234", "18908012345",
         "成都市锦江区人民南路一段97号锦江宾馆旁", "四川天府银行股份有限公司",
         "韩夫 刘波", "13908012345", "乳腺外科", "赵军", "乳腺结节", "SC510102198303120013"],
    ]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [16, 10, 6, 6, 22, 15, 40, 26, 14, 15, 14, 10, 20, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/医院患者登记表.xlsx")
    print("已生成: 医院患者登记表.xlsx")


def create_loan_application():
    """银行贷款申请"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "贷款申请"

    headers = ["申请编号", "申请人", "身份证号", "手机号", "邮箱",
               "家庭住址", "工作单位", "月收入(元)", "贷款金额(万元)", "贷款用途",
               "担保人", "担保人身份证", "担保人电话", "放款账号", "审批状态"]
    styled_header(ws, headers, "E65100")

    data = [
        ["LN-2024-001", "唐俊杰", "110105199208151234", "18610012345", "tangjunjie@gmail.com",
         "北京市海淀区知春路甲48号盈都大厦C座601", "北京字节跳动科技有限公司", "45000", "150",
         "购房首付", "唐父 唐建国", "110105196509121234", "13810012345",
         "6225881234512345678", "已批准"],
        ["LN-2024-002", "梁美玲", "440106199103151234", "15823456789", "liangmeiling@163.com",
         "广州市海珠区新港中路356号金逸华庭3栋1801", "广州市天河区第一小学", "12000", "30",
         "装修", "梁母 吴秀英", "440106196801151234", "13523456789",
         "6217001234509871234", "审核中"],
        ["LN-2024-003", "方志诚", "360102199507151234", "15879012345", "fangzhicheng@outlook.com",
         "南昌市东湖区八一大道369号万达星城B区7栋1602", "自由职业", "28000", "50",
         "创业经营", "赖明辉", "360102199301081234", "13870123456",
         "6217002345678901234", "已拒绝"],
        ["LN-2024-004", "沈雨薇", "320102199603081234", "15212345678", "shenyuwei@netease.com",
         "南京市鼓楼区中山北路405号恒基中心21楼", "南京网易有限公司", "32000", "80",
         "购车", "沈父 沈大明", "320102196801081234", "13312345678",
         "6228481234567890005", "已批准"],
        ["LN-2024-005", "谢国强", "440106198709231234", "13823456789", "xieguoqiang@139.com",
         "广州市天河区天河北路233号中信广场公寓16楼B座", "广州市越秀区融信金融服务有限公司", "38000", "200",
         "购房首付", "谢妻 周丹", "440106199001231234", "13723456780",
         "6222031234567890001", "审核中"],
    ]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [16, 10, 22, 15, 28, 40, 28, 14, 14, 12, 14, 22, 15, 22, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/银行贷款申请表.xlsx")
    print("已生成: 银行贷款申请表.xlsx")


def create_property_owner():
    """物业业主信息"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "业主信息"

    headers = ["房号", "业主姓名", "身份证号", "联系电话", "备用电话",
               "电子邮箱", "户籍地址", "工作单位", "车位号", "车牌号", "入住日期"]
    styled_header(ws, headers, "7B1FA2")

    data = [
        ["1-101", "范国庆", "110105196805121234", "13901234567", "010-65891234",
         "fanguoqing@126.com", "北京市东城区东四十条甲22号", "北京恒基物业管理有限公司",
         "B1-056", "京A·12345", "2020-03-15"],
        ["1-201", "秦晓丽", "110101199105081234", "13901234568", "",
         "qinxiaoli@hjwy.com", "北京市朝阳区朝阳公园南路12号", "北京恒基物业管理有限公司",
         "B1-057", "京A·23456", "2020-05-01"],
        ["2-1201", "唐俊杰", "110105199208151234", "18610012345", "",
         "tangjunjie@gmail.com", "北京市海淀区知春路甲48号", "北京字节跳动科技有限公司",
         "B2-123", "京Q·78901", "2024-12-01"],
        ["2-1502", "赵丽娜", "110108199312061234", "13611012345", "13711012345",
         "zhaolina@baidu.com", "北京市海淀区上地十街10号", "北京百度网讯科技有限公司",
         "B2-156", "京A·56789", "2021-08-20"],
        ["3-801", "钟嘉铭", "110101198801231234", "15010012345", "",
         "zhongjm@tencent.com", "广东省深圳市南山区高新科技园", "北京腾讯科技有限公司",
         "B3-088", "京B·34567", "2022-01-10"],
        ["3-1901", "马天宇", "110105199506151234", "18510012345", "010-88001234",
         "matianyu@meituan.com", "北京市朝阳区望京东路6号", "北京三快在线科技有限公司",
         "B3-190", "京Q·90123", "2023-06-01"],
        ["4-2301", "顾明远", "320505198803121234", "13812345678", "",
         "gumingyuan@xhzn.com", "江苏省苏州市工业园区星湖街218号", "苏州星河智能制造科技有限公司",
         "B4-230", "苏E·12345", "2024-01-15"],
    ]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [10, 10, 22, 15, 15, 26, 30, 28, 10, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/物业业主信息表.xlsx")
    print("已生成: 物业业主信息表.xlsx")


def create_school_register():
    """学校学生信息"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生信息"

    headers = ["学号", "姓名", "性别", "身份证号", "出生日期",
               "家庭住址", "父亲姓名", "父亲电话", "父亲单位",
               "母亲姓名", "母亲电话", "母亲单位", "紧急邮箱"]
    styled_header(ws, headers, "C62828")

    data = [
        ["2024010101", "钱嘉乐", "男", "110101201509151234", "2015-09-15",
         "北京市西城区月坛北街25号月坛雅居3号楼801", "钱伟明", "13801234560",
         "北京中信证券股份有限公司", "林小燕", "13901234560",
         "北京市西城区实验小学", "qianwm@citics.com"],
        ["2024010102", "孙思琪", "女", "110105201508201234", "2015-08-20",
         "北京市朝阳区亮马桥路50号世纪星源公寓5-302", "孙浩", "18923456789",
         "深圳市锦程房地产开发有限公司北京办事处", "周丹", "13723456780",
         "北京市朝阳区外国语学校", "sunhao@jincheng-re.com"],
        ["2024010103", "陈天佑", "男", "110108201507121234", "2015-07-12",
         "北京市海淀区清华东路甲35号院12号楼401", "陈国栋", "13512345678",
         "广东正义律师事务所北京分所", "赵启明", "13823456789",
         "清华大学附属小学", "chenguodong@gdzylawyer.com"],
        ["2024010104", "林思涵", "女", "110101201506081234", "2015-06-08",
         "北京市东城区东华门大街10号皇城根遗址公园旁", "林志远", "13607012345",
         "北京中伦律师事务所", "何思雨", "13723456789",
         "北京市东城区史家胡同小学", "linzhiyuan@nclaw.com"],
        ["2024010105", "韩子轩", "男", "110105201504231234", "2015-04-23",
         "北京市朝阳区朝外大街乙12号昆泰国际中心16层", "韩博", "13612345678",
         "深圳前海启航创业投资基金管理有限公司", "杜芸芸", "18712345678",
         "北京市朝阳区芳草地国际学校", "hanbo@qhqichuang.com"],
    ]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    col_widths = [14, 10, 6, 22, 14, 40, 10, 15, 30, 10, 15, 28, 28]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/学校学生信息登记表.xlsx")
    print("已生成: 学校学生信息登记表.xlsx")


if __name__ == "__main__":
    create_case_register()
    create_patient_register()
    create_loan_application()
    create_property_owner()
    create_school_register()
