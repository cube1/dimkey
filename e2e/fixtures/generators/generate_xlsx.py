#!/usr/bin/env python3
"""生成用于脱敏工具测试的 Excel 文件"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_employee_xlsx():
    """员工花名册 - 包含多种敏感信息混合在各列中"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工花名册"

    headers = ["工号", "姓名", "性别", "手机号", "身份证号", "邮箱",
               "银行卡号(工资卡)", "家庭住址", "紧急联系人", "紧急联系电话",
               "所属部门", "所属公司"]

    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    data = [
        ["EMP001", "张三", "男", "13800138001", "110101199003076789", "zhangsan@qq.com",
         "6222021234567890123", "北京市海淀区中关村大街1号院2号楼301室", "张母", "13800138002",
         "研发部", "北京星辰科技有限公司"],
        ["EMP002", "李四", "女", "15912345678", "320106198507121234", "lisi@163.com",
         "6217001234567890001", "上海市浦东新区张江高科技园区碧波路690号", "李父", "15912345679",
         "产品部", "上海云帆信息技术有限公司"],
        ["EMP003", "王五", "男", "18676543210", "440305199212253456", "wangwu@gmail.com",
         "6228480012345678901", "广东省深圳市南山区科技园南区高新南一道008号", "王妻", "18676543211",
         "市场部", "深圳前海智联科技有限公司"],
        ["EMP004", "赵六", "女", "13711112222", "510107199108150098", "zhaoliu@outlook.com",
         "6212261234567891234", "四川省成都市武侯区天府大道北段1700号", "赵母", "13711112223",
         "财务部", "成都天府数据服务有限公司"],
        ["EMP005", "陈七", "男", "17088889999", "330102198612305678", "chenqi@foxmail.com",
         "6225881234567890002", "浙江省杭州市西湖区文三路90号东部软件园", "陈父", "17088889998",
         "运维部", "杭州湖畔网络科技有限公司"],
        ["EMP006", "刘八", "女", "15233334444", "420106199305208765", "liuba@126.com",
         "6214830012345678003", "湖北省武汉市洪山区光谷大道77号金融港B11栋", "刘夫", "15233334445",
         "人力资源部", "武汉光谷创新信息技术有限公司"],
        ["EMP007", "孙九", "男", "18955556666", "340102199409100012", "sunjiu@sina.com",
         "6230580012345678904", "安徽省合肥市蜀山区望江西路800号创新产业园A区", "孙妻", "18955556667",
         "法务部", "合肥量子智能科技有限公司"],
        ["EMP008", "周十", "女", "13666667777", "500103199710251234", "zhoushi@hotmail.com",
         "6216610012345678905", "重庆市渝中区解放碑民权路28号英利国际金融中心", "周母", "13666667778",
         "行政部", "重庆山城大数据有限公司"],
        ["EMP009", "吴芳", "女", "14700001111", "230102199601080034", "wufang@yeah.net",
         "6222081234567890006", "黑龙江省哈尔滨市南岗区西大直街92号", "吴父", "14700001112",
         "研发部", "哈尔滨冰城软件有限公司"],
        ["EMP010", "郑刚", "男", "18300002222", "370102199802145678", "zhenggang@aliyun.com",
         "6217991234567890007", "山东省济南市历下区泉城路180号齐鲁国际大厦", "郑母", "18300002223",
         "测试部", "济南泉城信息科技有限公司"],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx <= 3 else 'left')

    # 调整列宽
    col_widths = [10, 10, 6, 15, 22, 25, 24, 40, 10, 15, 14, 28]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/员工花名册.xlsx")
    print("已生成: 员工花名册.xlsx")


def create_mixed_xlsx():
    """混合场景 - 敏感信息散布在非结构化文本中"""
    wb = openpyxl.Workbook()

    # Sheet1: 合同信息
    ws1 = wb.active
    ws1.title = "合同信息"
    ws1.append(["合同编号", "甲方信息", "乙方信息", "联系方式", "备注"])
    contracts = [
        ["HT-2024-001",
         "甲方: 北京星辰科技有限公司\n法定代表人: 张三\n地址: 北京市海淀区中关村大街1号",
         "乙方: 上海云帆信息技术有限公司\n法定代表人: 李四\n地址: 上海市浦东新区碧波路690号",
         "张三手机: 13800138001\n邮箱: zhangsan@qq.com\n李四手机: 15912345678",
         "合同金额50万，分三期支付。甲方银行账号: 6222021234567890123"],
        ["HT-2024-002",
         "甲方: 深圳前海智联科技有限公司\n法定代表人: 王五\n身份证号: 440305199212253456",
         "乙方: 杭州湖畔网络科技有限公司\n法定代表人: 陈七\n地址: 杭州市西湖区文三路90号",
         "王五电话: 18676543210\n陈七电话: 17088889999\n邮箱: wangwu@gmail.com",
         "合同有效期一年。乙方开户行: 工商银行杭州支行，卡号6225881234567890002"],
    ]
    for row in contracts:
        ws1.append(row)
    for col in range(1, 6):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 45

    # Sheet2: 客服工单
    ws2 = wb.create_sheet("客服工单")
    ws2.append(["工单号", "客户描述", "处理备注"])
    tickets = [
        ["WO-20240301-001",
         "客户马晓红(手机13501234567)反馈账户异常，身份证号210102198801156789，要求核实近期交易记录。",
         "已核实客户身份，通过邮箱maxiaohong@aliyun.com发送了交易流水。客户地址: 沈阳市沈河区青年大街185号。"],
        ["WO-20240301-002",
         "林建国先生来电(18298765432)咨询贷款事宜。客户提供身份证350102199202287654和工资卡号6217001234567890001。",
         "已转接至信贷部门处理，客户邮箱linjianguo@yeah.net。工作单位: 福州市鼓楼区五四路162号华城国际。"],
        ["WO-20240302-001",
         "用户黄丽萍投诉充值未到账，手机号15087654321。充值银行卡尾号8901，全卡号6228480012345678901。",
         "经查充值延迟到账，已补发到客户账户。客户确认收到，联系地址: 昆明市盘龙区东风东路36号。"],
    ]
    for row in tickets:
        ws2.append(row)
    for col in range(1, 4):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 60

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/混合敏感信息场景.xlsx")
    print("已生成: 混合敏感信息场景.xlsx")


def create_edge_case_xlsx():
    """边界测试 - 各种格式的敏感信息"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "边界测试"
    ws.append(["测试场景", "测试数据", "期望识别类型"])

    cases = [
        # 手机号各种格式
        ["手机号-标准格式", "13800138001", "Phone"],
        ["手机号-带区号", "+86 13800138001", "Phone"],
        ["手机号-带横线", "138-0013-8001", "Phone"],
        ["手机号-带空格", "138 0013 8001", "Phone"],
        ["手机号-嵌入文本", "请联系张经理13800138001或发邮件", "Phone"],
        ["手机号-座机", "010-62345678", "Phone"],
        # 身份证号
        ["身份证-18位", "110101199003076789", "IdCard"],
        ["身份证-末尾X", "11010119900307678X", "IdCard"],
        ["身份证-嵌入文本", "持证人身份证号码为110101199003076789请核实", "IdCard"],
        # 邮箱
        ["邮箱-标准", "zhangsan@qq.com", "Email"],
        ["邮箱-含数字", "zhang123@163.com", "Email"],
        ["邮箱-含下划线", "zhang_san@gmail.com", "Email"],
        ["邮箱-含点号", "zhang.san@company.com.cn", "Email"],
        ["邮箱-嵌入文本", "请发送至zhangsan@qq.com邮箱", "Email"],
        # 银行卡号
        ["银行卡-16位", "6222021234567890", "BankCard"],
        ["银行卡-19位", "6222021234567890123", "BankCard"],
        ["银行卡-带空格", "6222 0212 3456 7890 123", "BankCard"],
        # 姓名（需NER识别）
        ["姓名-两字", "张三", "PersonName"],
        ["姓名-三字", "欧阳修", "PersonName"],
        ["姓名-四字", "司马相如", "PersonName"],
        ["姓名-嵌入文本", "项目经理李明确认了方案", "PersonName"],
        # 地址
        ["地址-完整", "北京市海淀区中关村大街1号院2号楼301室", "Address"],
        ["地址-简短", "深圳市南山区科技园", "Address"],
        ["地址-含小区", "杭州市西湖区翠苑街道翠苑一区23幢401", "Address"],
        # 公司名
        ["公司-标准", "北京星辰科技有限公司", "Company"],
        ["公司-含括号", "腾讯科技(深圳)有限公司", "Company"],
        ["公司-集团", "中国平安保险(集团)股份有限公司", "Company"],
        # 多种敏感信息混合在一个单元格
        ["混合-短文本", "张三，手机13800138001，邮箱zhangsan@qq.com", "Phone,Email,PersonName"],
        ["混合-长文本",
         "客户王五(身份证440305199212253456)家住深圳市南山区科技园，联系电话18676543210，工资卡6228480012345678901，单位为深圳前海智联科技有限公司。",
         "IdCard,Address,Phone,BankCard,Company,PersonName"],
    ]
    for row in cases:
        ws.append(row)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 30

    wb.save("/Users/tanzeshun/workpath/git/desensitize-tool/test-data/边界测试用例.xlsx")
    print("已生成: 边界测试用例.xlsx")


if __name__ == "__main__":
    create_employee_xlsx()
    create_mixed_xlsx()
    create_edge_case_xlsx()
